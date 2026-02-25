from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Tuple

from relluna.core.document_memory import ConfidenceState, ProvenancedString

FONTE = "services.ocr.service"


@dataclass(frozen=True)
class OCROptions:
    enabled: bool = True
    min_text_len: int = 20           # abaixo disso, considera "texto insuficiente"
    max_pages: int = 3               # limite de páginas para OCR raster (evitar custo)
    dpi: int = 200                   # DPI para renderização no OCR raster


def _env_flag(name: str, default: str = "1") -> bool:
    import os
    val = os.getenv(name, default).strip().lower()
    return val in {"1", "true", "yes", "on"}


def get_ocr_options_from_env() -> OCROptions:
    import os

    enabled = _env_flag("RELLUNA_OCR_ENABLED", "1")
    min_text_len = int(os.getenv("RELLUNA_OCR_MIN_TEXT_LEN", "20"))
    max_pages = int(os.getenv("RELLUNA_OCR_MAX_PAGES", "3"))
    dpi = int(os.getenv("RELLUNA_OCR_DPI", "200"))
    return OCROptions(enabled=enabled, min_text_len=min_text_len, max_pages=max_pages, dpi=dpi)


def _safe_strip(txt: Optional[str]) -> str:
    return (txt or "").replace("\x00", "").strip()


def _preprocess_image_for_ocr(img):
    """
    Pré-processa imagem para melhorar OCR:
    - Redimensiona se muito pequena (mínimo 300 DPI equivalente)
    - Converte para grayscale
    - Aumenta contraste
    - Aplica nitidez
    """
    from PIL import ImageEnhance, ImageFilter
    
    # Converter para RGB se necessário
    if img.mode in ('RGBA', 'P'):
        img = img.convert('RGB')
    
    # Redimensionar se muito pequena (mínimo 1000px na maior dimensão)
    min_size = 1000
    if max(img.size) < min_size:
        ratio = min_size / max(img.size)
        new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
        img = img.resize(new_size, Image.LANCZOS)
    
    # Converter para grayscale
    img = img.convert('L')
    
    # Aumentar contraste
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    
    # Aplicar nitidez
    img = img.filter(ImageFilter.SHARPEN)
    
    return img


def _extract_pdf_text_native_pymupdf(path: Path, max_pages: int) -> Tuple[str, str]:
    """
    Extração nativa (PDF digital) via PyMuPDF (fitz).
    Retorna (texto, metodo).
    """
    import fitz  # PyMuPDF

    doc = fitz.open(str(path))
    parts = []
    n = min(len(doc), max_pages) if max_pages > 0 else len(doc)
    for i in range(n):
        page = doc.load_page(i)
        parts.append(page.get_text("text") or "")
    return "\n".join(parts), "pymupdf.get_text"


def _extract_pdf_text_native_pypdf(path: Path, max_pages: int) -> Tuple[str, str]:
    """
    Fallback nativo via pypdf (mais leve, mas em PDFs complexos pode vir vazio).
    """
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    n = min(len(reader.pages), max_pages) if max_pages > 0 else len(reader.pages)
    parts = []
    for i in range(n):
        try:
            parts.append(reader.pages[i].extract_text() or "")
        except Exception:
            parts.append("")
    return "\n".join(parts), "pypdf.extract_text"


def _ocr_image_tesseract(image_path: Path) -> Tuple[str, str]:
    """
    OCR raster via Tesseract (pytesseract). Requer:
      - apt-get install tesseract-ocr
      - pip install pytesseract
    """
    from PIL import Image
    import pytesseract

    with Image.open(image_path) as img:
        img.load()
        img = _preprocess_image_for_ocr(img)
        txt = pytesseract.image_to_string(img)
    return txt, "tesseract.image_to_string+preprocess"


def _ocr_pdf_raster_tesseract(path: Path, max_pages: int, dpi: int) -> Tuple[str, str]:
    """
    Renderiza páginas do PDF e roda OCR via tesseract.
    Requer PyMuPDF + pytesseract + tesseract-ocr.
    """
    import fitz  # PyMuPDF
    from PIL import Image
    import pytesseract
    import io

    doc = fitz.open(str(path))
    n = min(len(doc), max_pages) if max_pages > 0 else len(doc)

    parts = []
    for i in range(n):
        page = doc.load_page(i)
        mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
        pix = page.get_pixmap(matrix=mat, alpha=False)

        img = Image.open(io.BytesIO(pix.tobytes("png")))
        img = _preprocess_image_for_ocr(img)
        txt = pytesseract.image_to_string(img)
        parts.append(txt or "")

    return "\n".join(parts), "pymupdf.render + tesseract+preprocess"


def extract_text_for_path(path: Path, *, options: Optional[OCROptions] = None) -> Tuple[str, str]:
    options = options or get_ocr_options_from_env()

    if not options.enabled:
        return "", "disabled"

    if not path.exists():
        return "", "missing_file"

    suffix = path.suffix.lower()

    # ---------------- PDF ----------------
    if suffix == ".pdf":
        try:
            text, metodo = _extract_pdf_text_native_pymupdf(path, options.max_pages)
        except Exception:
            try:
                text, metodo = _extract_pdf_text_native_pypdf(path, options.max_pages)
            except Exception:
                text, metodo = "", "native_failed"

        text = _safe_strip(text)
        if len(text) >= options.min_text_len:
            return text, metodo

        try:
            text2, metodo2 = _ocr_pdf_raster_tesseract(path, options.max_pages, options.dpi)
            return _safe_strip(text2), metodo2
        except Exception:
            return text, metodo

    # ---------------- DOCX ----------------
    if suffix == ".docx":
        try:
            from docx import Document
            doc = Document(path)
            text = "\n".join(p.text for p in doc.paragraphs)
            return _safe_strip(text), "python-docx"
        except Exception:
            return "", "docx_failed"

    # ---------------- XLSX ----------------
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            parts = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(str(c) for c in row if c is not None)
                    if row_text:
                        parts.append(row_text)
            return _safe_strip("\n".join(parts)), "openpyxl"
        except Exception:
            return "", "xlsx_failed"

    # ---------------- CSV ----------------
    if suffix == ".csv":
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
            return _safe_strip(text), "csv_read"
        except Exception:
            return "", "csv_failed"

    # ---------------- TXT ----------------
    if suffix == ".txt":
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
            return _safe_strip(text), "txt_read"
        except Exception:
            return "", "txt_failed"

    # ---------------- IMAGEM ----------------
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"}:
        try:
            text, metodo = _ocr_image_tesseract(path)
            return _safe_strip(text), metodo
        except Exception:
            return "", "image_ocr_unavailable"

    return "", "unsupported"


def make_layer2_ocr_field(path: Path, *, options: Optional[OCROptions] = None) -> ProvenancedString:
    """
    Wrapper para gerar um ProvenancedString de texto OCR/literal para Layer2.

    - Usa extract_text_for_path para extrair texto.
    - Se não houver texto suficiente, marca como insuficiente (valor=None).
    - Mantém FONTE e metodo para rastreabilidade.
    """
    text, metodo = extract_text_for_path(path, options=options)
    text = _safe_strip(text)

    if not text:
        return ProvenancedString(
            valor=None,
            fonte=FONTE,
            metodo=metodo,
            estado=ConfidenceState.insuficiente,
            confianca=None,
        )

    return ProvenancedString(
        valor=text,
        fonte=FONTE,
        metodo=metodo,
        estado=ConfidenceState.confirmado,
        confianca=1.0,
    )
